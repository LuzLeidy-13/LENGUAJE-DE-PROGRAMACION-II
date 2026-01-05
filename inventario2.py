from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import json
import os
from tkinter import PhotoImage
from decimal import Decimal, ROUND_HALF_UP

# =============================
# MODELO MEJORADO (POO)
# =============================
class Producto:
    def __init__(self, codigo, nombre, descripcion="", precio=0.0, stock=0, stock_minimo=5, categoria="General", proveedor=""):
        self.codigo = codigo
        self.nombre = nombre
        self.descripcion = descripcion
        self.precio = Decimal(str(precio))
        self.stock = int(stock)
        self.stock_minimo = int(stock_minimo)
        self.categoria = categoria
        self.proveedor = proveedor
        self.fecha_creacion = datetime.now()
        self.fecha_actualizacion = datetime.now()
    
    def to_dict(self):
        return {
            'codigo': self.codigo,
            'nombre': self.nombre,
            'descripcion': self.descripcion,
            'precio': float(self.precio),
            'stock': self.stock,
            'stock_minimo': self.stock_minimo,
            'categoria': self.categoria,
            'proveedor': self.proveedor,
            'fecha_creacion': self.fecha_creacion.isoformat(),
            'fecha_actualizacion': self.fecha_actualizacion.isoformat()
        }
    
    @classmethod
    def from_dict(cls, data):
        producto = cls(
            data['codigo'],
            data['nombre'],
            data.get('descripcion', ''),
            data.get('precio', 0.0),
            data.get('stock', 0),
            data.get('stock_minimo', 5),
            data.get('categoria', 'General'),
            data.get('proveedor', '')
        )
        producto.fecha_creacion = datetime.fromisoformat(data['fecha_creacion'])
        producto.fecha_actualizacion = datetime.fromisoformat(data['fecha_actualizacion'])
        return producto
    
    def entrada_stock(self, cantidad, motivo=""):
        if cantidad > 0:
            self.stock += cantidad
            self.fecha_actualizacion = datetime.now()
            return True
        return False
    
    def salida_stock(self, cantidad, motivo=""):
        if 0 < cantidad <= self.stock:
            self.stock -= cantidad
            self.fecha_actualizacion = datetime.now()
            return True
        return False
    
    def valor_total(self):
        return self.precio * self.stock
    
    def necesita_reabastecer(self):
        return self.stock <= self.stock_minimo
    
    def __str__(self):
        return f"{self.codigo} - {self.nombre} (Stock: {self.stock})"


class Movimiento:
    def __init__(self, tipo, codigo_producto, cantidad, motivo="", usuario="admin"):
        self.fecha = datetime.now()
        self.tipo = tipo  # ENTRADA, SALIDA, AJUSTE
        self.codigo_producto = codigo_producto
        self.cantidad = cantidad
        self.motivo = motivo
        self.usuario = usuario
    
    def to_dict(self):
        return {
            'fecha': self.fecha.isoformat(),
            'tipo': self.tipo,
            'codigo_producto': self.codigo_producto,
            'cantidad': self.cantidad,
            'motivo': self.motivo,
            'usuario': self.usuario
        }


class Inventario:
    def __init__(self):
        self.productos = {}
        self.movimientos = []
        self.categorias = set(['General', 'Electrónica', 'Ropa', 'Alimentos', 'Librería', 'Hogar'])
        self.cargar_datos()
    
    # ========== OPERACIONES CRUD ==========
    def agregar_producto(self, producto):
        if producto.codigo in self.productos:
            raise ValueError(f"El producto con código {producto.codigo} ya existe")
        
        if producto.categoria not in self.categorias:
            self.categorias.add(producto.categoria)
        
        self.productos[producto.codigo] = producto
        self.guardar_datos()
        return True
    
    def actualizar_producto(self, codigo, **kwargs):
        if codigo not in self.productos:
            raise ValueError(f"Producto {codigo} no encontrado")
        
        producto = self.productos[codigo]
        for key, value in kwargs.items():
            if hasattr(producto, key):
                if key == 'precio':
                    value = Decimal(str(value))
                setattr(producto, key, value)
        
        producto.fecha_actualizacion = datetime.now()
        self.guardar_datos()
        return True
    
    def eliminar_producto(self, codigo):
        if codigo in self.productos:
            del self.productos[codigo]
            self.guardar_datos()
            return True
        return False
    
    def buscar_producto(self, criterio, valor):
        resultados = []
        for producto in self.productos.values():
            attr = getattr(producto, criterio, None)
            if attr and valor.lower() in str(attr).lower():
                resultados.append(producto)
        return resultados
    
    # ========== MOVIMIENTOS ==========
    def registrar_movimiento(self, tipo, codigo, cantidad, motivo=""):
        if codigo not in self.productos:
            raise ValueError(f"Producto {codigo} no encontrado")
        
        producto = self.productos[codigo]
        movimiento = Movimiento(tipo, codigo, cantidad, motivo)
        
        if tipo == "ENTRADA":
            producto.entrada_stock(cantidad, motivo)
        elif tipo == "SALIDA":
            if not producto.salida_stock(cantidad, motivo):
                raise ValueError("Stock insuficiente")
        elif tipo == "AJUSTE":
            producto.stock = cantidad
            producto.fecha_actualizacion = datetime.now()
        
        self.movimientos.append(movimiento)
        self.guardar_datos()
        return True
    
    # ========== REPORTES Y CONSULTAS ==========
    def productos_bajo_stock(self):
        return [p for p in self.productos.values() if p.necesita_reabastecer()]
    
    def valor_inventario_total(self):
        total = Decimal('0')
        for producto in self.productos.values():
            total += producto.valor_total()
        return total
    
    def productos_por_categoria(self):
        categorias = {}
        for producto in self.productos.values():
            if producto.categoria not in categorias:
                categorias[producto.categoria] = []
            categorias[producto.categoria].append(producto)
        return categorias
    
    def historial_movimientos(self, codigo=None):
        if codigo:
            return [m for m in self.movimientos if m.codigo_producto == codigo]
        return self.movimientos[-100:]  # Últimos 100 movimientos
    
    # ========== PERSISTENCIA ==========
    def guardar_datos(self):
        datos = {
            'productos': {k: v.to_dict() for k, v in self.productos.items()},
            'movimientos': [m.to_dict() for m in self.movimientos],
            'categorias': list(self.categorias)
        }
        with open('inventario_data.json', 'w', encoding='utf-8') as f:
            json.dump(datos, f, indent=2, ensure_ascii=False)
    
    def cargar_datos(self):
        if os.path.exists('inventario_data.json'):
            try:
                with open('inventario_data.json', 'r', encoding='utf-8') as f:
                    datos = json.load(f)
                
                for codigo, prod_data in datos['productos'].items():
                    self.productos[codigo] = Producto.from_dict(prod_data)
                
                for mov_data in datos.get('movimientos', []):
                    movimiento = Movimiento(
                        mov_data['tipo'],
                        mov_data['codigo_producto'],
                        mov_data['cantidad'],
                        mov_data.get('motivo', ''),
                        mov_data.get('usuario', 'admin')
                    )
                    movimiento.fecha = datetime.fromisoformat(mov_data['fecha'])
                    self.movimientos.append(movimiento)
                
                self.categorias = set(datos.get('categorias', ['General']))
            except Exception as e:
                print(f"Error cargando datos: {e}")
    
    # ========== EXPORTACIONES ==========
    def exportar_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Inventario"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), 
                               right=Side(style='thin'),
                               top=Side(style='thin'), 
                               bottom=Side(style='thin'))
            
            # Encabezados
            headers = ["Código", "Nombre", "Descripción", "Categoría", "Precio", "Stock", 
                      "Stock Mínimo", "Valor Total", "Proveedor", "Última Actualización"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
            
            # Datos
            row = 2
            for producto in self.productos.values():
                ws.cell(row=row, column=1, value=producto.codigo).border = thin_border
                ws.cell(row=row, column=2, value=producto.nombre).border = thin_border
                ws.cell(row=row, column=3, value=producto.descripcion).border = thin_border
                ws.cell(row=row, column=4, value=producto.categoria).border = thin_border
                ws.cell(row=row, column=5, value=float(producto.precio)).border = thin_border
                ws.cell(row=row, column=6, value=producto.stock).border = thin_border
                ws.cell(row=row, column=7, value=producto.stock_minimo).border = thin_border
                ws.cell(row=row, column=8, value=float(producto.valor_total())).border = thin_border
                ws.cell(row=row, column=9, value=producto.proveedor).border = thin_border
                ws.cell(row=row, column=10, value=producto.fecha_actualizacion.strftime("%Y-%m-%d %H:%M")).border = thin_border
                
                # Resaltar bajo stock
                if producto.necesita_reabastecer():
                    for col in range(1, 11):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="FF9999", end_color="FF9999", fill_type="solid")
                row += 1
            
            # Ajustar ancho columnas
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 20
            
            # Hoja de resumen
            ws2 = wb.create_sheet("Resumen")
            ws2.append(["RESUMEN DEL INVENTARIO"])
            ws2.append(["Total Productos:", len(self.productos)])
            ws2.append(["Valor Total Inventario:", f"S/ {self.valor_inventario_total():,.2f}"])
            ws2.append(["Productos con Bajo Stock:", len(self.productos_bajo_stock())])
            
            # Guardar
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"inventario_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            )
            
            if filename:
                wb.save(filename)
                return filename
            return None
            
        except ImportError:
            messagebox.showerror("Error", "Instale openpyxl: pip install openpyxl")
            return None
    
    def generar_reporte_pdf(self, reporte_tipo="inventario"):
        try:
            from reportlab.lib.pagesizes import A4, letter, landscape
            from reportlab.pdfgen import canvas
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            from reportlab.lib.units import inch, cm
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                initialfile=f"reporte_{reporte_tipo}_{datetime.now().strftime('%Y%m%d')}.pdf"
            )
            
            if not filename:
                return None
            
            doc = SimpleDocTemplate(filename, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1
            )
            elements.append(Paragraph(f"REPORTE DE INVENTARIO - {datetime.now().strftime('%d/%m/%Y')}", title_style))
            
            if reporte_tipo == "inventario":
                # Datos para tabla
                data = [["Código", "Nombre", "Categoría", "Precio", "Stock", "Valor"]]
                for producto in self.productos.values():
                    data.append([
                        producto.codigo,
                        producto.nombre[:20],
                        producto.categoria,
                        f"S/ {float(producto.precio):.2f}",
                        str(producto.stock),
                        f"S/ {float(producto.valor_total()):,.2f}"
                    ])
                
                # Crear tabla
                table = Table(data, colWidths=[2*cm, 4*cm, 3*cm, 3*cm, 2*cm, 4*cm])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                ]))
                
                elements.append(table)
                elements.append(Spacer(1, 20))
                
                # Totales
                elements.append(Paragraph(f"Total Productos: {len(self.productos)}", styles['Normal']))
                elements.append(Paragraph(f"Valor Total Inventario: S/ {float(self.valor_inventario_total()):,.2f}", styles['Normal']))
            
            elif reporte_tipo == "bajo_stock":
                productos_bajos = self.productos_bajo_stock()
                if productos_bajos:
                    data = [["Código", "Nombre", "Stock Actual", "Stock Mínimo", "Diferencia"]]
                    for producto in productos_bajos:
                        data.append([
                            producto.codigo,
                            producto.nombre[:20],
                            str(producto.stock),
                            str(producto.stock_minimo),
                            str(producto.stock_minimo - producto.stock)
                        ])
                    
                    table = Table(data)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.red),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ]))
                    elements.append(table)
                else:
                    elements.append(Paragraph("No hay productos con bajo stock", styles['Normal']))
            
            doc.build(elements)
            return filename
            
        except ImportError:
            messagebox.showerror("Error", "Instale reportlab: pip install reportlab")
            return None


# =============================
# INTERFAZ MEJORADA
# =============================
class InventarioGUI:
    def __init__(self, root):
        self.inv = Inventario()
        self.root = root
        root.title("Sistema de Control de Inventarios - Pro")
        root.geometry("1200x700")
        
        # Configurar tema
        self.setup_styles()
        
        # Frame principal
        main_frame = ttk.Frame(root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Título
        title_frame = ttk.Frame(main_frame)
        title_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(title_frame, text="🏬 SISTEMA DE CONTROL DE INVENTARIOS", 
                 font=("Arial", 18, "bold"), foreground="#2c3e50").pack(side=tk.LEFT)
        
        # Frame de estadísticas
        self.stats_frame = ttk.LabelFrame(main_frame, text="Estadísticas")
        self.stats_frame.pack(fill=tk.X, pady=(0, 10))
        self.update_stats()
        
        # Frame de búsqueda
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(search_frame, text="Buscar:").pack(side=tk.LEFT, padx=(0, 5))
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=30)
        self.search_entry.pack(side=tk.LEFT, padx=(0, 10))
        self.search_entry.bind('<KeyRelease>', self.buscar_productos)
        
        ttk.Label(search_frame, text="Por:").pack(side=tk.LEFT, padx=(0, 5))
        self.search_criteria = ttk.Combobox(search_frame, values=["nombre", "código", "categoría", "proveedor"], width=15)
        self.search_criteria.set("nombre")
        self.search_criteria.pack(side=tk.LEFT, padx=(0, 10))
        
        # Frame principal dividido
        paned = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True)
        
        # Frame izquierdo - Formulario
        left_frame = ttk.Frame(paned)
        self.formulario_producto(left_frame)
        paned.add(left_frame, weight=1)
        
        # Frame derecho - Tabla y controles
        right_frame = ttk.Frame(paned)
        self.tabla_productos(right_frame)
        self.controles_avanzados(right_frame)
        paned.add(right_frame, weight=2)
        
        # Barra de estado
        self.status_bar = ttk.Label(root, text="Listo", relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Cargar datos iniciales
        self.refrescar_tabla()
    
    def setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        
        # Colores personalizados
        style.configure("TButton", padding=6)
        style.configure("Primary.TButton", background="#3498db", foreground="white")
        style.configure("Success.TButton", background="#2ecc71", foreground="white")
        style.configure("Warning.TButton", background="#e74c3c", foreground="white")
        style.configure("Info.TButton", background="#9b59b6", foreground="white")
    
    def update_stats(self):
        for widget in self.stats_frame.winfo_children():
            widget.destroy()
        
        stats = [
            ("📦 Total Productos", len(self.inv.productos)),
            ("💰 Valor Inventario", f"S/ {float(self.inv.valor_inventario_total()):,.2f}"),
            ("⚠️ Bajo Stock", len(self.inv.productos_bajo_stock())),
            ("📊 Categorías", len(self.inv.categorias))
        ]
        
        for i, (label, value) in enumerate(stats):
            frame = ttk.Frame(self.stats_frame)
            frame.grid(row=0, column=i, padx=20, pady=5)
            
            ttk.Label(frame, text=label, font=("Arial", 9)).pack()
            ttk.Label(frame, text=str(value), font=("Arial", 14, "bold"), 
                     foreground="#2c3e50").pack()
    
    def formulario_producto(self, parent):
        form_frame = ttk.LabelFrame(parent, text="Gestión de Productos", padding=10)
        form_frame.pack(fill=tk.BOTH, expand=True)
        
        # Campos del formulario
        fields = [
            ("Código:", "e_codigo"),
            ("Nombre:", "e_nombre"),
            ("Descripción:", "e_descripcion"),
            ("Categoría:", "e_categoria"),
            ("Precio (S/):", "e_precio"),
            ("Stock Inicial:", "e_stock"),
            ("Stock Mínimo:", "e_stock_min"),
            ("Proveedor:", "e_proveedor")
        ]
        
        self.form_vars = {}
        for i, (label, var_name) in enumerate(fields):
            ttk.Label(form_frame, text=label).grid(row=i, column=0, sticky=tk.W, pady=5)
            
            if var_name == "e_categoria":
                self.e_categoria = ttk.Combobox(form_frame, values=sorted(self.inv.categorias))
                self.e_categoria.grid(row=i, column=1, sticky=tk.EW, pady=5, padx=(5, 0))
                self.e_categoria.set("General")
            elif var_name == "e_descripcion":
                self.e_descripcion = tk.Text(form_frame, height=3, width=30)
                self.e_descripcion.grid(row=i, column=1, sticky=tk.EW, pady=5, padx=(5, 0))
            else:
                var = tk.StringVar()
                entry = ttk.Entry(form_frame, textvariable=var)
                entry.grid(row=i, column=1, sticky=tk.EW, pady=5, padx=(5, 0))
                setattr(self, var_name, entry)
                self.form_vars[var_name] = var
        
        # Botones del formulario
        btn_frame = ttk.Frame(form_frame)
        btn_frame.grid(row=len(fields), column=0, columnspan=2, pady=20)
        
        ttk.Button(btn_frame, text="➕ Agregar", command=self.agregar_producto,
                  style="Success.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="✏️ Actualizar", command=self.actualizar_producto,
                  style="Primary.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="🗑️ Eliminar", command=self.eliminar_producto,
                  style="Warning.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="🔄 Limpiar", command=self.limpiar_formulario).pack(side=tk.LEFT, padx=5)
        
        # Movimientos
        mov_frame = ttk.LabelFrame(form_frame, text="Movimientos de Stock", padding=10)
        mov_frame.grid(row=len(fields)+1, column=0, columnspan=2, sticky=tk.EW, pady=10)
        
        ttk.Label(mov_frame, text="Cantidad:").grid(row=0, column=0, padx=5)
        self.e_cantidad = ttk.Entry(mov_frame, width=10)
        self.e_cantidad.grid(row=0, column=1, padx=5)
        
        ttk.Label(mov_frame, text="Motivo:").grid(row=0, column=2, padx=5)
        self.e_motivo = ttk.Entry(mov_frame, width=20)
        self.e_motivo.grid(row=0, column=3, padx=5)
        
        ttk.Button(mov_frame, text="📥 Entrada", command=lambda: self.movimiento("ENTRADA"),
                  style="Success.TButton").grid(row=0, column=4, padx=5)
        ttk.Button(mov_frame, text="📤 Salida", command=lambda: self.movimiento("SALIDA"),
                  style="Warning.TButton").grid(row=0, column=5, padx=5)
        ttk.Button(mov_frame, text="⚙️ Ajuste", command=lambda: self.movimiento("AJUSTE"),
                  style="Info.TButton").grid(row=0, column=6, padx=5)
    
    def tabla_productos(self, parent):
        # Frame para la tabla
        table_frame = ttk.Frame(parent)
        table_frame.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbars
        vscroll = ttk.Scrollbar(table_frame)
        vscroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        hscroll = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL)
        hscroll.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Treeview con más columnas
        columns = ("Código", "Nombre", "Categoría", "Precio", "Stock", "Stock Mín", "Valor", "Proveedor")
        self.tv = ttk.Treeview(table_frame, columns=columns, show="headings",
                              yscrollcommand=vscroll.set, xscrollcommand=hscroll.set,
                              selectmode="browse", height=15)
        
        # Configurar columnas
        col_widths = [80, 150, 100, 80, 60, 80, 100, 120]
        for col, width in zip(columns, col_widths):
            self.tv.heading(col, text=col)
            self.tv.column(col, width=width, anchor="center")
        
        self.tv.pack(fill=tk.BOTH, expand=True)
        vscroll.config(command=self.tv.yview)
        hscroll.config(command=self.tv.xview)
        
        # Bind para selección
        self.tv.bind('<<TreeviewSelect>>', self.seleccionar_producto)
        
        # Etiqueta de información
        self.info_label = ttk.Label(parent, text="Doble clic para ver detalles")
        self.info_label.pack(pady=5)
    
    def controles_avanzados(self, parent):
        # Frame para botones avanzados
        adv_frame = ttk.LabelFrame(parent, text="Herramientas Avanzadas", padding=10)
        adv_frame.pack(fill=tk.X, pady=(10, 0))
        
        # Primera fila de botones
        row1 = ttk.Frame(adv_frame)
        row1.pack(fill=tk.X, pady=5)
        
        ttk.Button(row1, text="📊 Reporte Completo", command=lambda: self.generar_reporte("inventario"),
                  style="Primary.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(row1, text="⚠️ Reporte Bajo Stock", command=lambda: self.generar_reporte("bajo_stock"),
                  style="Warning.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(row1, text="📈 Historial Movimientos", command=self.ver_historial,
                  style="Info.TButton").pack(side=tk.LEFT, padx=5)
        
        # Segunda fila de botones
        row2 = ttk.Frame(adv_frame)
        row2.pack(fill=tk.X, pady=5)
        
        ttk.Button(row2, text="💾 Exportar Excel", command=self.exportar_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(row2, text="🔄 Actualizar Categorías", command=self.actualizar_categorias).pack(side=tk.LEFT, padx=5)
        ttk.Button(row2, text="📋 Copiar Datos", command=self.copiar_datos).pack(side=tk.LEFT, padx=5)
        ttk.Button(row2, text="🔄 Refrescar", command=self.refrescar_tabla).pack(side=tk.LEFT, padx=5)
    
    # ========== MÉTODOS DE ACCIÓN ==========
    def agregar_producto(self):
        try:
            # Validar campos obligatorios
            if not self.e_codigo.get() or not self.e_nombre.get():
                messagebox.showwarning("Validación", "Código y Nombre son obligatorios")
                return
            
            producto = Producto(
                codigo=self.e_codigo.get().upper(),
                nombre=self.e_nombre.get(),
                descripcion=self.e_descripcion.get("1.0", tk.END).strip(),
                precio=float(self.e_precio.get() or 0),
                stock=int(self.e_stock.get() or 0),
                stock_minimo=int(self.e_stock_min.get() or 5),
                categoria=self.e_categoria.get(),
                proveedor=self.e_proveedor.get()
            )
            
            self.inv.agregar_producto(producto)
            self.refrescar_tabla()
            self.limpiar_formulario()
            self.update_stats()
            self.status_bar.config(text=f"Producto {producto.codigo} agregado exitosamente")
            
        except ValueError as e:
            messagebox.showerror("Error", str(e))
        except Exception as e:
            messagebox.showerror("Error", f"Error inesperado: {str(e)}")
    
    def actualizar_producto(self):
        seleccion = self.tv.selection()
        if not seleccion:
            messagebox.showwarning("Seleccionar", "Seleccione un producto de la tabla")
            return
        
        codigo = self.tv.item(seleccion)['values'][0]
        
        try:
            datos = {
                'nombre': self.e_nombre.get(),
                'descripcion': self.e_descripcion.get("1.0", tk.END).strip(),
                'precio': float(self.e_precio.get() or 0),
                'stock': int(self.e_stock.get() or 0),
                'stock_minimo': int(self.e_stock_min.get() or 5),
                'categoria': self.e_categoria.get(),
                'proveedor': self.e_proveedor.get()
            }
            
            self.inv.actualizar_producto(codigo, **datos)
            self.refrescar_tabla()
            self.update_stats()
            self.status_bar.config(text=f"Producto {codigo} actualizado")
            
        except Exception as e:
            messagebox.showerror("Error", str(e))
    
    def eliminar_producto(self):
        seleccion = self.tv.selection()
        if not seleccion:
            messagebox.showwarning("Seleccionar", "Seleccione un producto de la tabla")
            return
        
        codigo = self.tv.item(seleccion)['values'][0]
        
        if messagebox.askyesno("Confirmar", f"¿Eliminar producto {codigo}?"):
            if self.inv.eliminar_producto(codigo):
                self.refrescar_tabla()
                self.limpiar_formulario()
                self.update_stats()
                self.status_bar.config(text=f"Producto {codigo} eliminado")
            else:
                messagebox.showerror("Error", "No se pudo eliminar el producto")
    
    def movimiento(self, tipo):
        seleccion = self.tv.selection()
        if not seleccion:
            messagebox.showwarning("Seleccionar", "Seleccione un producto de la tabla")
            return
        
        codigo = self.tv.item(seleccion)['values'][0]
        cantidad = self.e_cantidad.get()
        motivo = self.e_motivo.get()
        
        if not cantidad or not cantidad.isdigit():
            messagebox.showwarning("Validación", "Ingrese una cantidad válida")
            return
        
        cantidad = int(cantidad)
        
        try:
            self.inv.registrar_movimiento(tipo, codigo, cantidad, motivo)
            self.refrescar_tabla()
            self.update_stats()
            self.e_cantidad.delete(0, tk.END)
            self.e_motivo.delete(0, tk.END)
            self.status_bar.config(text=f"Movimiento {tipo} registrado para {codigo}")
            
        except Exception as e:
            messagebox.showerror("Error", str(e))
    
    def seleccionar_producto(self, event=None):
        seleccion = self.tv.selection()
        if seleccion:
            datos = self.tv.item(seleccion)['values']
            if datos:
                # Buscar producto completo
                producto = self.inv.productos.get(datos[0])
                if producto:
                    # Actualizar formulario
                    self.limpiar_formulario()
                    self.e_codigo.insert(0, producto.codigo)
                    self.e_nombre.insert(0, producto.nombre)
                    self.e_descripcion.insert("1.0", producto.descripcion)
                    self.e_categoria.set(producto.categoria)
                    self.e_precio.insert(0, str(producto.precio))
                    self.e_stock.insert(0, str(producto.stock))
                    self.e_stock_min.insert(0, str(producto.stock_minimo))
                    self.e_proveedor.insert(0, producto.proveedor)
                    
                    # Deshabilitar código para edición
                    self.e_codigo.config(state='readonly')
    
    def buscar_productos(self, event=None):
        criterio = self.search_criteria.get()
        valor = self.search_var.get()
        
        if not valor:
            self.refrescar_tabla()
            return
        
        resultados = self.inv.buscar_producto(criterio, valor)
        
        # Limpiar tabla
        for item in self.tv.get_children():
            self.tv.delete(item)
        
        # Mostrar resultados
        for producto in resultados:
            self.tv.insert("", "end", values=(
                producto.codigo,
                producto.nombre,
                producto.categoria,
                f"S/ {float(producto.precio):.2f}",
                producto.stock,
                producto.stock_minimo,
                f"S/ {float(producto.valor_total()):,.2f}",
                producto.prove
